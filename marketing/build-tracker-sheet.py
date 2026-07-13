#!/usr/bin/env python3
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

MD = "/home/user/build-a-pitcher/marketing/creator-targets.md"
OUT = "/tmp/claude-0/-home-user-build-a-pitcher/a1ac5da3-249b-5637-a4a7-4e5cec613fa4/scratchpad/GoatLab-Creator-Tracker.xlsx"

# section number -> (Sport, Segment)
SEC = {
    "1": ("Baseball", "Show / Gaming (YT+Twitch)"),
    "2": ("Baseball", "TikTok / Instagram"),
    "3": ("Baseball", "X / Twitter"),
    "4": ("Football", "College (culture+gaming)"),
    "5": ("Baseball", "Round 2 (mixed)"),
    "6": ("Basketball", "NBA / NBA 2K"),
    "7": ("Soccer", "EA FC / Football Manager"),
    "8": ("Hockey", "NHL / EA NHL"),
    "9": ("Football", "Round 2 (Madden/CFB/NFL)"),
    "10": ("Multi-sport", "Owner picks"),
}

rows = []
cur = None
with open(MD, encoding="utf-8") as f:
    for line in f:
        m = re.match(r"^##\s+(\d+)\.", line)
        if m:
            cur = m.group(1)
            continue
        if line.startswith("##"):
            # non-numeric section (first wave, outreach log, contents) -> stop tagging
            if not re.match(r"^##\s+\d+\.", line):
                cur = None
            continue
        if cur and line.startswith("|"):
            cells = [c.strip().replace("**", "").replace("`", "").strip() for c in line.strip().strip("|").split("|")]
            if len(cells) != 7:
                continue
            if cells[0].lower() == "creator" or cells[6].lower() == "priority" or set(cells[0]) <= set("-: "):
                continue
            creator, handle, followers, style, whyfit, rate, priority = cells
            sport, segment = SEC[cur]
            rows.append({
                "Priority": priority, "Sport": sport, "Segment": segment,
                "Creator": creator, "Platform & Handle": handle, "Followers": followers,
                "Style": style, "Why fit": whyfit, "Est. rate": rate,
            })

# sort: priority (High>Medium>Low), then sport
prank = {"High": 0, "Medium": 1, "Low": 2}
def pkey(p):
    p = p.lower()
    if "high" in p: return 0
    if "medium" in p: return 1
    return 2
rows.sort(key=lambda r: (pkey(r["Priority"]), r["Sport"], r["Creator"].lower()))

wb = Workbook()

# ---------- palette ----------
NAVY = "0A1320"; BLUE = "14304A"; CYAN = "19C6FF"
HEAD = PatternFill("solid", fgColor=BLUE)
BAND = PatternFill("solid", fgColor="F2F8FC")
HIGH = PatternFill("solid", fgColor="E4F7EC")
LOW = PatternFill("solid", fgColor="F3F4F6")
white = Font(color="FFFFFF", bold=True, size=11)
thin = Side(style="thin", color="D5E0EA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical="top", wrap_text=True)
top = Alignment(vertical="top")

# ============ TAB 1: Creators tracker ============
ws = wb.active
ws.title = "Creators"

info_cols = ["Priority", "Sport", "Segment", "Creator", "Platform & Handle", "Followers", "Style", "Why fit", "Est. rate"]
track_cols = ["Ref code", "Outreach date", "Channel", "Replied?", "Posted?", "Post type", "New users", "Paid?", "Notes"]
cols = info_cols + track_cols
widths = [9, 11, 22, 22, 30, 14, 26, 34, 16,  10, 13, 9, 9, 9, 12, 10, 8, 26]

# title row
ws.merge_cells("A1:R1")
t = ws["A1"]; t.value = "GoatLab — Creator Outreach Tracker"
t.font = Font(color="FFFFFF", bold=True, size=15); t.fill = PatternFill("solid", fgColor=NAVY)
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.row_dimensions[1].height = 30
ws.merge_cells("A2:R2")
s = ws["A2"]; s.value = ("Sort/filter by Priority or Sport. Fill the right-hand columns as you go. "
                         "Tip: select the Replied?/Posted?/Paid? columns and Insert > Checkbox in Google Sheets.")
s.font = Font(italic=True, color="44586C", size=10); s.alignment = Alignment(vertical="center", indent=1)
ws.row_dimensions[2].height = 22

hdr = 3
for j, c in enumerate(cols, 1):
    cell = ws.cell(hdr, j, c); cell.fill = HEAD; cell.font = white
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(j)].width = widths[j-1]
ws.row_dimensions[hdr].height = 26

for i, r in enumerate(rows):
    rr = hdr + 1 + i
    vals = [r[c] for c in info_cols] + [""] * len(track_cols)
    for j, v in enumerate(vals, 1):
        cell = ws.cell(rr, j, v)
        cell.border = border
        cell.alignment = wrap if j in (3,4,5,7,8,9,18) else top
        if j <= len(info_cols):
            pk = pkey(r["Priority"])
            cell.fill = HIGH if pk == 0 else (BAND if i % 2 else PatternFill())
    # priority cell emphasis
    ws.cell(rr, 1).font = Font(bold=True,
        color=("1E7D46" if pkey(r["Priority"])==0 else ("9A6B00" if pkey(r["Priority"])==1 else "6B7280")))

ws.freeze_panes = "D4"
ws.auto_filter.ref = f"A{hdr}:R{hdr+len(rows)}"

# dropdowns for Channel / Replied / Posted / Paid / Post type
def dv(colletter, formula):
    d = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(d)
    d.add(f"{colletter}{hdr+1}:{colletter}{hdr+len(rows)}")
dv("L", '"X DM,Email,IG DM,TikTok DM,Other"')       # Channel
dv("M", '"Yes,No,Waiting"')                          # Replied
dv("N", '"Yes,No,Scheduled"')                        # Posted
dv("O", '"Video,Post,Livestream,Story,Thread"')      # Post type
dv("Q", '"Yes,No,Partial"')                          # Paid

# ============ TAB 2: Pay & Rules ============
def sheet_block(ws, title, blocks):
    ws.merge_cells("A1:D1")
    c = ws["A1"]; c.value = title; c.font = Font(bold=True, size=15, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100
    row = 3
    for kind, text in blocks:
        cell = ws.cell(row, 2, text)
        if kind == "h":
            cell.font = Font(bold=True, size=12, color="0D2438")
        elif kind == "b":
            cell.value = "•  " + text
            cell.font = Font(size=11, color="1A2230")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            cell.font = Font(size=11, color="1A2230")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

pay = wb.create_sheet("Pay & Rules")
sheet_block(pay, "Pay & Ground Rules", [
    ("h", "Pay"),
    ("b", "$50/week for the first two weeks — a trial run for both of us."),
    ("b", "$10 bonus every time a creator you reached out to posts a video/post."),
    ("b", "$25 bonus for a livestream."),
    ("b", "$20 for every 1,000 new users your creators send us (tracked through the links — mostly YouTube, X, Twitch)."),
    ("b", "All bonuses capped at $500 total."),
    ("b", "Paid every Friday (Venmo/Zelle)."),
    ("b", "If it's clicking after two weeks, we keep going and bump the base."),
    ("p", ""),
    ("h", "How a bonus counts"),
    ("p", "A post only counts once it's LIVE and has the tracking link in it. A shoutout with no link "
          "doesn't count — we can't track it and it barely sends anyone."),
    ("p", ""),
    ("h", "Ground rules"),
    ("b", "Never say we're affiliated with MLB, MLB The Show, or EA. It's \"real MLB ratings,\" that's it."),
    ("b", "If anyone asks, be upfront you're reaching out on behalf of GoatLab."),
    ("b", "Keep it human in the DMs — a real person who likes the game, not a bot."),
    ("b", "Max ~10-15 reach-outs a day per platform. One follow-up only if they ghost (5-7 days later)."),
    ("b", "Personalize the first line every time — mention a recent post/video. No copy-paste blasts."),
    ("p", ""),
    ("h", "Who to target"),
    ("b", "All sports (baseball, basketball, soccer, football) are fair game; gaming creators are a secondary lane."),
    ("b", "At least 5k followers, active in the last couple weeks, real engagement (not bots)."),
    ("b", "Work the list on the Creators tab top-down by Priority, and add your own finds to the bottom."),
])

# ============ TAB 3: Templates ============
tpl = wb.create_sheet("Message Templates")
sheet_block(tpl, "Outreach Templates (personalize the [brackets])", [
    ("h", "1. Cold DM — TikTok / Instagram / X (free seeding)"),
    ("p", "Hey [name] — [one specific line about a recent post of theirs]. I made a free browser game I think "
          "your audience would eat up: spin a random real [sport] player, choose where each rating goes on his "
          "body, then simulate his whole career. No download, takes 3 minutes: goat-lab.app/?ref=[code]"),
    ("p", "No ask — just thought it was your kind of thing. If you ever want to post it, the funniest format we've "
          "seen is \"building the worst player possible\" and reading the career verdict."),
    ("p", ""),
    ("h", "Follow-up (once, 5-7 days later, only if no reply)"),
    ("p", "One more nudge and I'll leave you alone — today's Daily Challenge gives everyone the same cards, so "
          "\"beat my score\" actually works on your audience: goat-lab.app/?ref=[code]"),
    ("p", ""),
    ("h", "2. Email — YouTube / Twitch (paid)"),
    ("p", "Subject: Sponsor idea — build-a-99 browser game your audience will get instantly"),
    ("p", "Hi [name], I'm with GoatLab (goat-lab.app), a free browser game where you spin real players with real "
          "ratings, assign each to a body part, build a 99 OVR, and simulate the career. I'd love to sponsor a video."),
    ("p", "Formats that work: God Squad vs Cursed Squad (best vs worst build), a 1v1 vs a viewer, or \"Beat my "
          "Daily\" where your audience plays the same cards."),
    ("p", "Offer: $[X] for one dedicated video (or $[Y] integration + a pinned-comment Daily follow-up a week "
          "later). You'd use your tracked link goat-lab.app/?ref=[code] and I'll share the click-to-play numbers. "
          "Creative is all yours; only ask is the link in the description and saying it's free. Interested?"),
    ("p", ""),
    ("h", "3. Paid-offer terms (after they say yes)"),
    ("b", "Deliverable: [1 dedicated video / 2 posts a week apart / 1 stream segment 20+ min]"),
    ("b", "Fee: $[X], paid [50% upfront / on posting] via [Venmo/PayPal]"),
    ("b", "Required: tracked link goat-lab.app/?ref=[code] in [description/bio/pinned comment]; mention it's free"),
    ("b", "Disclosure: mark it #ad/sponsored (FTC)"),
    ("b", "Creative control fully theirs; no exclusivity"),
])

# ============ TAB 4: How to use ============
how = wb.create_sheet("How to use")
sheet_block(how, "How to use this sheet", [
    ("h", "Get it into Google Sheets"),
    ("b", "Upload this file to Google Drive, right-click > Open with > Google Sheets (it converts, keeping all tabs)."),
    ("b", "Or in a blank Google Sheet: File > Import > Upload > this file."),
    ("p", ""),
    ("h", "Working the Creators tab"),
    ("b", "Sort or filter by Priority (High first) or Sport using the filter arrows on the header row."),
    ("b", "For each creator you contact, fill: Ref code, Outreach date, Channel, then Replied?/Posted?/Paid? as it happens."),
    ("b", "Want real checkboxes? Select the Replied?/Posted?/Paid? columns, then Insert > Checkbox."),
    ("b", "New users = the number from the tracking-link report (ask [owner] for it); it drives the $20/1,000 bonus."),
    ("b", "Add creators you source yourself to the bottom rows — handle, follower count, sport, and one line on why they fit."),
    ("p", ""),
    ("h", "Ref codes"),
    ("b", "Each creator gets a short lowercase code (e.g. koogs). [Owner] generates the link goat-lab.app/?ref=<code>."),
    ("b", "The link only tracks clicks that turn into plays, so it must actually be IN the post (bio/description)."),
])

for w in (pay, tpl, how):
    for rr in range(1, w.max_row + 1):
        w.row_dimensions[rr].height = max(w.row_dimensions[rr].height or 15, 15)

wb.save(OUT)
print(f"wrote {OUT} with {len(rows)} creators across {len(wb.sheetnames)} tabs: {wb.sheetnames}")
